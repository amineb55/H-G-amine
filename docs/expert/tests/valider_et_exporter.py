"""
Validation croisée des livrables experts + export Excel pour relecture.

Usage (depuis la racine du dépôt) :
    python docs/expert/tests/valider_et_exporter.py

- Charge exigences/*.yaml, regles_terrain/*.yaml, documents_types.yaml,
  formulaire, glossaire, baremes, guard.
- Vérifie : ids uniques, champs obligatoires, slugs de documents connus,
  ids de règles terrain connus (BTP-/BUR- tolérés : fichiers existants du dépôt),
  garde anti-copyright (motifs interdits), poids 1-3, type valide.
- Exporte des classeurs Excel avec colonnes de relecture (Validation / Commentaire).
Sortie non nulle si une erreur bloquante est trouvée.
"""
from __future__ import annotations
import glob, os, sys, re
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.environ.get("EXPERT_XLSX_OUT", ROOT)
TYPES_OK = {"documentaire", "mise_en_oeuvre", "enregistrement"}
CHAMPS = ["id", "chapitre", "titre_fr", "titre_en", "exigence_fr", "exigence_en", "type",
          "preuves_attendues", "questions_audit", "documents_types", "hls_commun",
          "equivalences", "applicabilite", "regles_terrain", "erreurs_frequentes", "poids"]
errors, warnings = [], []

def load(path):
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)

# ---------- chargement ----------
exigences = []
for p in sorted(glob.glob(os.path.join(ROOT, "exigences", "exigences_*.yaml"))):
    exigences += load(p)
doc_types = load(os.path.join(ROOT, "exigences", "documents_types.yaml"))
regles = {}
for p in sorted(glob.glob(os.path.join(ROOT, "regles_terrain", "*.yaml"))):
    d = load(p)
    for r in d["regles"]:
        regles[r["id"]] = r | {"secteur": d["secteur"]}
guard = load(os.path.join(ROOT, "guard", "motifs_interdits.yaml"))
formulaire = load(os.path.join(ROOT, "formulaire", "questions_cadrage.yaml"))
glossaire = load(os.path.join(ROOT, "glossaire", "glossaire_fr_en.yaml"))
baremes = load(os.path.join(ROOT, "baremes", "baremes.yaml"))
load(os.path.join(ROOT, "outils_qualite", "arbre_decision.yaml"))

# ---------- validation exigences ----------
ids = set()
for e in exigences:
    eid = e.get("id", "?")
    for c in CHAMPS:
        if c not in e:
            errors.append(f"{eid}: champ manquant '{c}'")
    if eid in ids:
        errors.append(f"{eid}: id en double")
    ids.add(eid)
    if e.get("type") not in TYPES_OK:
        errors.append(f"{eid}: type invalide {e.get('type')}")
    if e.get("poids") not in (1, 2, 3):
        errors.append(f"{eid}: poids invalide {e.get('poids')}")
    for d in e.get("documents_types", []):
        if d not in doc_types:
            errors.append(f"{eid}: type de document inconnu '{d}'")
    for r in e.get("regles_terrain", []):
        if r not in regles and not r.startswith(("BTP-", "BUR-")):
            errors.append(f"{eid}: règle terrain inconnue '{r}'")
    txt = e.get("exigence_fr", "")
    if not (250 <= len(txt) <= 900):
        warnings.append(f"{eid}: exigence_fr de {len(txt)} caractères (cible 250-900)")
    for m in guard["interdits_fr"]:
        if m in txt:
            errors.append(f"{eid}: motif interdit « {m} »")
    for m in guard["interdits_en"]:
        if m in e.get("exigence_en", ""):
            errors.append(f"{eid}: motif interdit EN « {m} »")
    if re.search(r"\b(doit|doivent)\b", txt):
        errors.append(f"{eid}: verbe d'obligation « doit/doivent » dans exigence_fr")
    for m in guard["a_eviter_fr"]:
        if m in txt:
            warnings.append(f"{eid}: motif à éviter « {m} »")
    if not e.get("preuves_attendues") or not e.get("questions_audit"):
        errors.append(f"{eid}: preuves ou questions vides")
    extra = set(e) - set(CHAMPS) - {"nuance_de"}
    if extra:
        errors.append(f"{eid}: champ(s) inconnu(s) {sorted(extra)}")

# références croisées entre normes (après chargement complet)
for e in exigences:
    eid = e["id"]
    if e.get("nuance_de") and e["nuance_de"] not in ids:
        errors.append(f"{eid}: nuance_de pointe vers un id inconnu '{e['nuance_de']}'")
    for x in e.get("equivalences", []):
        if x not in ids:
            warnings.append(f"{eid}: équivalence {x} sans enregistrement propre (exigence HLS portée par le tronc commun 45001)")

# ---------- validation règles terrain ----------
for rid, r in regles.items():
    for c in ("titre", "categorie", "description", "criticite_defaut", "exigences", "indices_visuels", "exclusions"):
        if c not in r:
            errors.append(f"{rid}: champ manquant '{c}'")
    if r.get("criticite_defaut") not in (1, 2, 3, 4):
        errors.append(f"{rid}: criticite_defaut invalide")
    for x in r.get("exigences", []):
        if x not in ids:
            errors.append(f"{rid}: exigence inconnue '{x}'")
    if len(r.get("indices_visuels", [])) < 3:
        warnings.append(f"{rid}: moins de 3 indices visuels")
    if len(r.get("exclusions", [])) < 2:
        warnings.append(f"{rid}: moins de 2 exclusions")

# ---------- validation formulaire ----------
qids = set()
for q in formulaire["questions"]:
    if q["id"] in qids:
        errors.append(f"{q['id']}: question en double")
    qids.add(q["id"])
    for d in q.get("documents", []):
        if d not in doc_types:
            errors.append(f"{q['id']}: type de document inconnu '{d}'")
    for x in q.get("exigences", []):
        if x not in ids and not x.startswith(("9001-", "14001-", "50001-")):
            errors.append(f"{q['id']}: exigence inconnue '{x}'")
        elif x not in ids:
            warnings.append(f"{q['id']}: exigence {x} en attente du modèle correspondant (apport 2)")

# ---------- export Excel ----------
FONT = Font(name="Arial", size=10)
HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F4E78")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")

def sheet(wb, title, headers, rows, widths, review=True, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    hdr = headers + (["Validation Amine (OK / à modifier / à supprimer)", "Commentaire Amine"] if review else [])
    ws.append(hdr)
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=i); c.font = HEAD; c.fill = FILL
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for r in rows:
        ws.append(list(r) + (["", ""] if review else []))
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = FONT; c.alignment = Alignment(wrap_text=True, vertical="top")
        if review:
            row[-1].fill = REVIEW_FILL; row[-2].fill = REVIEW_FILL
    for i, w in enumerate(widths + ([22, 40] if review else []), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"
    return ws

def lisezmoi(wb, lignes):
    ws = wb.active; ws.title = "Lisez-moi"
    for l in lignes:
        ws.append([l])
    for row in ws.iter_rows():
        for c in row:
            c.font = FONT; c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120
    ws["A1"].font = Font(name="Arial", size=12, bold=True)

def j(lst):  # liste -> texte multi-lignes
    return "\n".join(f"- {x}" for x in lst) if lst else ""

# Exigences
wb = Workbook()
lisezmoi(wb, [
    "MODÈLES D'EXIGENCES ISO 45001 / 9001 / 14001 / 50001 — BROUILLON v0.1 à relire par Amine Boukha",
    "Source de vérité : docs/expert/exigences/exigences_*.yaml (ce classeur est généré depuis les YAML). Un onglet par norme.",
    "Tronc commun HLS : les exigences communes sont portées par les enregistrements 45001 marqués hls_commun = oui ; les onglets 9001 / 14001 / 50001 ne contiennent que les exigences spécifiques et les nuances (colonne nuance_de).",
    "ISO 50001 : à faire valider par un auditeur énergie qualifié avant mise en production.",
    "Colonnes jaunes = à remplir par le relecteur : Validation (OK / à modifier / à supprimer) et Commentaire. Le partenaire de structuration reporte ensuite les corrections dans les YAML.",
    "Règle de relecture : chaque exigence_fr décrit l'état constaté par l'auditeur, jamais la phrase de la norme ; si une phrase vous semble reconnaissable à l'identique, réécrivez-la.",
    "Exemple de ligne validée : id 45001-6.1.2.1 (enregistrement de référence fourni dans le cahier des charges) → Validation = OK.",
    "Nombre d'enregistrements : " + ", ".join(f"{n} = {sum(1 for e in exigences if e['id'].startswith(n + '-'))}" for n in ("45001", "9001", "14001", "50001")) + ".",
])
for norme in ("45001", "9001", "14001", "50001"):
    sub = [e for e in exigences if e["id"].startswith(norme + "-")]
    sheet(wb, f"Exigences {norme}",
          ["id", "chapitre", "titre_fr", "titre_en", "exigence_fr", "exigence_en", "type", "preuves_attendues",
           "questions_audit", "documents_types", "hls_commun", "equivalences", "nuance_de", "applicabilite",
           "regles_terrain", "erreurs_frequentes", "poids"],
          [(e["id"], e["chapitre"], e["titre_fr"], e["titre_en"], e["exigence_fr"].strip(), e["exigence_en"], e["type"],
            j(e["preuves_attendues"]), j(e["questions_audit"]), ", ".join(e["documents_types"]), e["hls_commun"],
            ", ".join(e["equivalences"]), e.get("nuance_de", ""), e["applicabilite"], ", ".join(e["regles_terrain"]),
            e["erreurs_frequentes"].strip(), e["poids"]) for e in sub],
          [13, 8, 28, 26, 70, 30, 14, 45, 45, 30, 9, 20, 13, 25, 22, 45, 6])
sheet(wb, "Types de documents", ["slug", "libellé", "famille"],
      [(k, v["libelle"], v["famille"]) for k, v in doc_types.items()], [34, 60, 16], review=False)
wb.save(os.path.join(OUT, "exigences", "exigences_45001_relecture.xlsx"))

# Règles terrain
wb = Workbook()
lisezmoi(wb, [
    "RÉFÉRENTIELS TERRAIN — industrie, santé, logistique, hôtellerie-restauration — BROUILLON v0.1",
    "Source de vérité : docs/expert/regles_terrain/*.yaml (classeur généré). Colonnes jaunes à remplir par le relecteur.",
    "Les champs les plus importants pour le moteur de détection sont indices_visuels (ce que l'image montre) et exclusions (ce qui ressemble à une NC sans en être une).",
    "criticite_defaut : 4 critique, 3 majeure, 2 mineure, 1 observation (voir baremes/baremes.yaml pour les escalades).",
    "Exemple de ligne validée : IND-09 issue de secours obstruée → Validation = OK.",
    f"Nombre de règles : {len(regles)}.",
])
sheet(wb, "Règles terrain",
      ["id", "secteur", "titre", "categorie", "description", "criticite_defaut", "exigences",
       "reference_reglementaire", "indices_visuels", "exclusions", "action_immediate"],
      [(r["id"], r["secteur"], r["titre"], r["categorie"], r["description"].strip(), r["criticite_defaut"],
        ", ".join(r["exigences"]), r.get("reference_reglementaire", ""), j(r["indices_visuels"]),
        j(r["exclusions"]), r.get("action_immediate", "")) for r in regles.values()],
      [9, 12, 34, 14, 45, 9, 28, 32, 60, 50, 40])
wb.save(os.path.join(OUT, "regles_terrain", "regles_terrain_relecture.xlsx"))

# Formulaire
wb = Workbook()
lisezmoi(wb, [
    "FORMULAIRE DE CADRAGE — questions — BROUILLON v0.1",
    "Source de vérité : docs/expert/formulaire/questions_cadrage.yaml (classeur généré). Colonnes jaunes à remplir par le relecteur.",
    "seuil_minimal = oui : réponse obligatoire pour lancer la génération v1 ; secteur/norme = filtres d'affichage.",
    "Exemple de ligne validée : Q01 raison sociale et sites → Validation = OK.",
    "Sections : " + " | ".join(f"{k}. {v}" for k, v in formulaire["sections"].items()),
])
sheet(wb, "Questions",
      ["id", "section", "secteur", "norme", "seuil_minimal", "type_reponse", "question", "documents alimentés",
       "exigences alimentées", "pourquoi (aide affichée)"],
      [(q["id"], q["section"], q["secteur"], q["norme"], q["seuil_minimal"], q["type_reponse"], q["question"],
        ", ".join(q["documents"]), ", ".join(q["exigences"]), q["pourquoi"]) for q in formulaire["questions"]],
      [7, 8, 11, 8, 12, 16, 70, 34, 28, 50])
wb.save(os.path.join(OUT, "formulaire", "questions_cadrage_relecture.xlsx"))

# Glossaire
wb = Workbook()
lisezmoi(wb, [
    "GLOSSAIRE FR / EN — BROUILLON v0.1",
    "Source de vérité : docs/expert/glossaire/glossaire_fr_en.yaml (classeur généré). Colonnes jaunes à remplir par le relecteur.",
    "Définitions rédigées en propres mots (pas les définitions ISO) ; ajustez les traductions à votre usage.",
    "Exemple de ligne validée : « Presqu'accident / Near miss » → Validation = OK.",
    f"Nombre de termes : {len(glossaire['termes'])}.",
])
sheet(wb, "Glossaire", ["terme_fr", "terme_en", "domaine", "definition_fr"],
      [(t["fr"], t["en"], t["dom"], t["def"]) for t in glossaire["termes"]], [36, 36, 14, 90])
wb.save(os.path.join(OUT, "glossaire", "glossaire_fr_en_relecture.xlsx"))

# ---------- bilan ----------
chap = {}
for e in exigences:
    chap[e["chapitre"]] = chap.get(e["chapitre"], 0) + 1
for n in ("45001", "9001", "14001", "50001"):
    print(f"Exigences {n} : {sum(1 for e in exigences if e['id'].startswith(n + '-'))}")
print(f"Total exigences : {len(exigences)} (par chapitre : {dict(sorted(chap.items()))})")
print(f"Types de documents : {len(doc_types)} | Règles terrain : {len(regles)} | Questions : {len(formulaire['questions'])} | Termes : {len(glossaire['termes'])}")
print(f"Barèmes : niveaux={len(baremes['criticite_constats']['niveaux'])}, degrés couverture={len(baremes['couverture_documentaire']['degres'])}")
for w in warnings:
    print("AVERTISSEMENT:", w)
for e in errors:
    print("ERREUR:", e)
print(f"{len(errors)} erreur(s), {len(warnings)} avertissement(s)")
sys.exit(1 if errors else 0)
